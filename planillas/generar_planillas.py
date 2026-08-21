# -*- coding: utf-8 -*-
"""
Genera el libro de Control de Stock Semanal de Candela Cafe & Patisserie.

El catalogo de areas, productos e insumos NO se tipea a mano: se extrae del
propio index.html de la app de Pedidos de Produccion, de modo que si el
catalogo cambia alcanza con volver a correr este script.

Uso:
    pip install openpyxl
    python3 generar_planillas.py [ruta/a/index.html] [salida.xlsx]

Requiere node en el PATH (solo para evaluar los literales del catalogo).
"""
import json, os, subprocess, sys, tempfile, unicodedata

# ---------------------------------------------------------------------------
# 1) Extraccion del catalogo desde index.html
# ---------------------------------------------------------------------------
INICIO = 'const PRODUCTOS_DEFAULT'
FIN    = 'const ADMIN_PASS'

def extraer_catalogo(ruta_html):
    with open(ruta_html, encoding='utf-8') as fh:
        html = fh.read()
    i, j = html.find(INICIO), html.find(FIN)
    if i < 0 or j < 0 or j <= i:
        raise SystemExit('No se encontro el bloque de catalogo en %s '
                         '(se buscaba entre "%s" y "%s").' % (ruta_html, INICIO, FIN))
    js = html[i:j] + """
require('fs').writeFileSync(process.argv[2], JSON.stringify({
  AREAS: AREAS,
  PRODUCTOS: PRODUCTOS_DEFAULT,
  INSUMOS: INSUMOS_PRODUCCION,
  INSUMOS_VENTAS: INSUMOS_VENTAS
}));
"""
    tmp = tempfile.mkdtemp()
    js_path, out_path = os.path.join(tmp, 'cat.js'), os.path.join(tmp, 'cat.json')
    with open(js_path, 'w', encoding='utf-8') as fh:
        fh.write(js)
    try:
        subprocess.run(['node', js_path, out_path], check=True, capture_output=True)
    except FileNotFoundError:
        raise SystemExit('Hace falta node en el PATH para leer el catalogo de index.html.')
    except subprocess.CalledProcessError as e:
        raise SystemExit('node no pudo evaluar el catalogo:\n' + e.stderr.decode('utf-8', 'replace'))
    with open(out_path, encoding='utf-8') as fh:
        return json.load(fh)

AQUI     = os.path.dirname(os.path.abspath(__file__))
RUTA_HTML = sys.argv[1] if len(sys.argv) > 1 else os.path.join(AQUI, '..', 'index.html')
SALIDA    = sys.argv[2] if len(sys.argv) > 2 else os.path.join(AQUI, 'Control_Stock_Semanal_Candela.xlsx')
CAT = extraer_catalogo(RUTA_HTML)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter as CL


# ---------------- paleta (tomada de css/styles.css de la app) ----------------
F      = 'Arial'
PRIM   = '6B3A2A'   # --color-primary
PRIMDK = '4E2A1E'   # --color-primary-dk
ACC    = 'C8963E'   # --color-accent
ACCLT  = 'F0C97A'   # --color-accent-lt
BG     = 'FFF8F0'   # --color-bg
BORD   = 'E8D5C0'   # --color-border
TXT    = '2C1A0E'
MUTED  = '7A5C4A'
WARNBG = 'FFF3CD'
DANGBG = 'FDECEA'
OKBG   = 'EAF7ED'
INFOBG = 'E3F0FF'

IN_FILL  = PatternFill('solid', fgColor=WARNBG)   # celda a completar
CALCFILL = PatternFill('solid', fgColor='FFFFFF')
BANDFILL = PatternFill('solid', fgColor=BG)
HDRFILL  = PatternFill('solid', fgColor=PRIMDK)
SUBFILL  = PatternFill('solid', fgColor=PRIM)
ACCFILL  = PatternFill('solid', fgColor=ACCLT)
TOTFILL  = PatternFill('solid', fgColor=BORD)

def fnt(sz=10, b=False, color=TXT, it=False): return Font(name=F, size=sz, bold=b, color=color, italic=it)
FIN   = fnt(10, color='0000FF')            # input manual (azul)
FLINK = fnt(10, color='008000')            # link a otra hoja (verde)
FCALC = fnt(10)                            # formula (negro)

thin = Side(style='thin', color=BORD)
med  = Side(style='medium', color=PRIM)
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM  = '#,##0.00;[Red]-#,##0.00;-'
NUM0 = '#,##0;[Red]-#,##0;-'
MON  = '$#,##0.00;[Red]-$#,##0.00;-'
PCT  = '0.0%;[Red]-0.0%;-'
DATE = 'dd/mm/yyyy'

# ---------------- helpers ----------------
def put(ws, r, c, v, font=None, fill=None, fmt=None, align=None, wrap=False,
        box=True, hor=None, ver='center'):
    cel = ws.cell(row=r, column=c, value=v)
    cel.font = font or FCALC
    if fill: cel.fill = fill
    if fmt:  cel.number_format = fmt
    if box:  cel.border = BOX
    cel.alignment = Alignment(horizontal=hor or align, vertical=ver, wrap_text=wrap)
    return cel

def titulo(ws, ncols, t, sub=None, sub2=None):
    """Bloque de titulo: fila1 titulo, fila2 subtitulo, fila3 subtitulo 2."""
    last = CL(ncols)
    ws.merge_cells(f'A1:{last}1')
    c = ws['A1']; c.value = t
    c.font = fnt(14, True, 'FFFFFF'); c.fill = HDRFILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 30
    if sub is not None:
        ws.merge_cells(f'A2:{last}2')
        c = ws['A2']; c.value = sub
        c.font = fnt(9, False, 'FFFFFF'); c.fill = SUBFILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 18
    if sub2 is not None:
        ws.merge_cells(f'A3:{last}3')
        c = ws['A3']; c.value = sub2
        c.font = fnt(9, True, PRIMDK); c.fill = ACCFILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[3].height = 18

def cabecera(ws, row, headers, heights=28):
    for i, h in enumerate(headers, start=1):
        c = put(ws, row, i, h, font=fnt(9, True, 'FFFFFF'), fill=SUBFILL,
                hor='center', wrap=True)
    ws.row_dimensions[row].height = heights

def anchos(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[CL(i)].width = w

def imprimir(ws, ncols, nrows, landscape=True, repeat='1:4'):
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = repeat
    ws.print_area = f'A1:{CL(ncols)}{nrows}'
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.oddFooter.right.text = "Pag. &P de &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.left.text = "Candela - Control de Stock Semanal"
    ws.oddFooter.left.size = 8

def bandear(ws, r0, r1, ncols):
    for r in range(r0, r1 + 1):
        if (r - r0) % 2 == 1:
            for c in range(1, ncols + 1):
                cel = ws.cell(row=r, column=c)
                try:
                    tiene = cel.fill is not None and cel.fill.patternType is not None
                except Exception:
                    tiene = False
                if not tiene:
                    cel.fill = BANDFILL

def norm(s):
    return ''.join(ch for ch in unicodedata.normalize('NFD', s)
                   if unicodedata.category(ch) != 'Mn').lower()

# ---------------- datos: areas, insumos, productos ----------------
AREAS = CAT['AREAS']
ALAB  = {a['key']: a['label'] for a in AREAS}
# Produccion: 'C' = obrador central, 'L' = se elabora en el local
DONDE = {'panaderia':'Central','pasteleria':'Central','especialidades':'Central',
         'factureria':'Central','sandwiches':'Central','cocina_sl':'Local San Luis'}

# unidad base + contenido por bulto, inferidos del nombre del insumo.
def unidad_bulto(nombre):
    n = norm(nombre)
    if 'x25kg' in n:                     return 'kg', 25
    if 'x kg' in n:                      return 'kg', 1
    if 'x unidad' in n:                  return 'u', 1
    if 'x caja' in n:                    return 'caja', 1
    if n.startswith('huevos'):           return 'u', 30
    if 'aceite' in n:                    return 'L', 5
    if 'lavandina' in n or 'detergente' in n: return 'L', 5
    if 'leche entera' in n:              return 'L', 1
    if 'crema de leche' in n:            return 'L', 1
    if 'crema vegetal' in n:             return 'L', 1
    if 'esencia' in n or 'variegato' in n: return 'L', 1
    if 'miel' in n:                      return 'kg', 5
    if any(k in n for k in ('bandeja','bolsas','vasos','servilletas','sorbetes')): return 'u', 1
    return 'kg', 1

# insumos que arrancan marcados como CRITICOS (alto costo / mayor riesgo)
CRITICOS = {norm(x) for x in [
    'Harina 000 x25kg','Harina 0000 x25kg','Margarina hojaldre','Margarina masa','Grasa',
    'Manteca','Azucar','Huevos','Levadura','Chocolate moldeo','Chocolate gotas',
    'Chocolate chip','Bano lodiser semiamargo','Baño lodiser semiamargo','Baño alpino',
    'Dulce de leche repostero','Crema de leche','Crema Vegetal','Nutella','Pistachos',
    'Nueces','Almendras','Leche entera','Queso crema','Queso cremoso','Queso sardo',
    'Queso tybo','Jamon cocido','Queso de máquina barra','Queso cheddar fetas',
    'Queso parmesano rallo','Queso roquefort','Jamón cocido paladini','Salame milán',
    'Lomo ahumado feteado','Panceta ahumada feteada','Palta x kg','Frutilla x kg',
    'Arándanos x caja','Mix de frutos rojos','Frutos del bosque','Arandanos congelados',
]}

INS = []   # (area_label, nombre, categoria, proveedor, unidad, bulto, critico)
for a in AREAS:
    for it in CAT['INSUMOS'].get(a['key'], []):
        u, b = unidad_bulto(it['name'])
        INS.append((ALAB[a['key']], it['name'], it.get('cat', ''), it.get('prov', ''),
                    u, b, 'SI' if norm(it['name']) in CRITICOS else 'NO'))

PROD = {a['key']: CAT['PRODUCTOS'].get(a['key'], []) for a in AREAS}
EXTRA = {k: 8 for k in PROD}
EXTRA['cocina_sl'] = 40          # Cocina San Luis no tiene catalogo cargado

NINS = len(INS)
IR0, IR1 = 5, 4 + NINS           # filas de datos de insumos (identicas en todas las hojas)

DIAS = ['LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO','DOMINGO']
MOTIVOS = ['Quemado / mal cocido','Mal armado / defectuoso','Vencido / pasado',
           'Caido / roto','Sobrante no vendido','Prueba / degustacion',
           'Consumo del personal','Devolucion de local','Otro']
UNIDADES = ['kg','u','L','caja','docena','bandeja']

wb = Workbook()
wb.remove(wb.active)

# =========================================================================
# Creacion de hojas en orden de pestanas
# =========================================================================
NOMS = ['Instructivo','Config','Insumos','Productos','1 Inv Inicial','2 Ingresos',
        '3 Conteo Diario']
PRODSH = {}
for a in AREAS:
    nm = '4 Prod ' + ''.join(ch for ch in unicodedata.normalize('NFD', a['label'])
                             if unicodedata.category(ch) != 'Mn').replace('Cocina San Luis','Cocina SL')
    PRODSH[a['key']] = nm
    NOMS.append(nm)
NOMS += ['5 Descartes','6 Inv Final','7 Facturas','8 Fudo Ventas','8 Fudo Stock',
         '8 Fudo Consumos','A Consumo Insumos','B Concil Productos','C Consumo Diario',
         'Tablero','Listas']
SH = {n: wb.create_sheet(n) for n in NOMS}

CFG_D = ['Config!$B$%d' % (10 + i) for i in range(7)]     # fechas Lun..Dom

# =========================================================================
# Hoja LISTAS (oculta) — origen de los desplegables
# =========================================================================
ws = SH['Listas']
ws['A1'] = 'AREAS'; ws['B1'] = 'INSUMOS (unicos)'; ws['C1'] = 'MOTIVOS'
ws['D1'] = 'UNIDADES'; ws['E1'] = 'PRODUCTOS (unicos)'; ws['F1'] = 'TIPO CONSUMO'
for c in 'ABCDEF': ws[c + '1'].font = fnt(10, True)
for i, a in enumerate(AREAS): ws.cell(row=2 + i, column=1, value=ALAB[a['key']])
uins = sorted({x[1] for x in INS})
for i, v in enumerate(uins): ws.cell(row=2 + i, column=2, value=v)
for i, v in enumerate(MOTIVOS): ws.cell(row=2 + i, column=3, value=v)
for i, v in enumerate(UNIDADES): ws.cell(row=2 + i, column=4, value=v)
uprod = sorted({p['name'] for a in AREAS for p in PROD[a['key']]})
for i, v in enumerate(uprod): ws.cell(row=2 + i, column=5, value=v)
for i, v in enumerate(['Consumo del personal','Cortesia','Prueba / degustacion','Merma','Otro']):
    ws.cell(row=2 + i, column=6, value=v)
ws.sheet_state = 'hidden'

DV_AREA = lambda: DataValidation(type='list', formula1='Listas!$A$2:$A$%d' % (1 + len(AREAS)),
                                 allow_blank=True, showErrorMessage=False)
DV_INS  = lambda: DataValidation(type='list', formula1='Listas!$B$2:$B$%d' % (1 + len(uins)),
                                 allow_blank=True, showErrorMessage=False)
DV_MOT  = lambda: DataValidation(type='list', formula1='Listas!$C$2:$C$%d' % (1 + len(MOTIVOS)),
                                 allow_blank=True, showErrorMessage=False)
DV_UNI  = lambda: DataValidation(type='list', formula1='Listas!$D$2:$D$%d' % (1 + len(UNIDADES)),
                                 allow_blank=True, showErrorMessage=False)
DV_PROD = lambda: DataValidation(type='list', formula1='Listas!$E$2:$E$%d' % (1 + len(uprod)),
                                 allow_blank=True, showErrorMessage=False)
DV_TIPO = lambda: DataValidation(type='list', formula1='Listas!$F$2:$F$6',
                                 allow_blank=True, showErrorMessage=False)
DV_SN   = lambda: DataValidation(type='list', formula1='"SI,NO"',
                                 allow_blank=True, showErrorMessage=False)

# =========================================================================
# Hoja CONFIG
# =========================================================================
ws = SH['Config']
anchos(ws, [34, 24, 26, 26])
titulo(ws, 4, 'CONFIGURACION DE LA SEMANA DE CONTROL',
       'Candela Cafe & Patisserie  ·  Complete SOLO las celdas amarillas. Todo el libro se actualiza solo.',
       'Esta hoja alimenta las fechas y los desplegables del resto de las planillas.')

def seccion(ws, r, txt, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=txt)
    c.font = fnt(11, True, 'FFFFFF'); c.fill = SUBFILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r].height = 22

seccion(ws, 5, '1 · SEMANA DE CONTROL', 4)
put(ws, 6, 1, 'Fecha de inicio (lunes)', fnt(10, True))
put(ws, 6, 2, None, FIN, IN_FILL, DATE, hor='center')
put(ws, 6, 3, 'Escriba la fecha del lunes en que arranca el control.', fnt(9, it=True, color=MUTED), box=False)
put(ws, 7, 1, 'Fecha de cierre (domingo)', fnt(10, True))
put(ws, 7, 2, '=IF($B$6="","",$B$6+6)', FCALC, None, DATE, hor='center')
put(ws, 7, 3, 'Se calcula solo.', fnt(9, it=True, color=MUTED), box=False)

cabecera(ws, 9, ['DIA', 'FECHA', 'OBSERVACIONES DEL DIA', ''], 20)
for i, d in enumerate(DIAS):
    r = 10 + i
    put(ws, r, 1, d.capitalize(), fnt(10, True), BANDFILL if i % 2 else None)
    put(ws, r, 2, '=IF($B$6="","",$B$6+%d)' % i, FCALC, None, DATE, hor='center')
    put(ws, r, 3, None, FIN, IN_FILL)
    put(ws, r, 4, None, FIN, IN_FILL)
ws.merge_cells('C9:D9')

seccion(ws, 18, '2 · RESPONSABLES POR AREA', 4)
cabecera(ws, 19, ['AREA', 'DONDE SE ELABORA', 'JEFE DE AREA (nombre)', 'FIRMA / ACLARACION'], 22)
for i, a in enumerate(AREAS):
    r = 20 + i
    put(ws, r, 1, ALAB[a['key']], fnt(10, True), BANDFILL if i % 2 else None)
    put(ws, r, 2, DONDE[a['key']], FIN, IN_FILL, hor='center')
    put(ws, r, 3, None, FIN, IN_FILL)
    put(ws, r, 4, None, FIN, IN_FILL)

seccion(ws, 27, '3 · MOTIVOS DE DESCARTE (desplegable en la hoja "5 Descartes")', 4)
for i, m in enumerate(MOTIVOS):
    r = 28 + i
    put(ws, r, 1, m, fnt(10), BANDFILL if i % 2 else None)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    put(ws, r, 3, None, box=False)

seccion(ws, 38, '4 · LEYENDA DE COLORES Y TIPOS DE LETRA', 4)
LEY = [('Celda amarilla, letra azul', 'La completa una persona a mano. Es el unico lugar donde hay que escribir.'),
       ('Letra negra, fondo blanco', 'Formula. NO tocar: se calcula sola.'),
       ('Letra verde', 'Trae el dato de otra hoja del libro. NO tocar.'),
       ('Fondo rojo claro', 'Alerta: el desvio supera el limite tolerado. Hay que investigar.'),
       ('Fondo verde claro', 'Dentro de lo esperado.')]
for i, (k, v) in enumerate(LEY):
    r = 39 + i
    put(ws, r, 1, k, fnt(10, True))
    put(ws, r, 2, v, fnt(10), wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 16
ws['A39'].fill = IN_FILL; ws['A39'].font = FIN
ws['A41'].font = FLINK
ws['A42'].fill = PatternFill('solid', fgColor=DANGBG)
ws['A43'].fill = PatternFill('solid', fgColor=OKBG)

seccion(ws, 45, '5 · LIMITES DE TOLERANCIA (para las alertas del Tablero)', 4)
put(ws, 46, 1, 'Desvio maximo tolerado en insumos (%)', fnt(10, True))
put(ws, 46, 2, 0.05, FIN, IN_FILL, PCT, hor='center')
put(ws, 46, 3, 'Diferencia entre consumo real y consumo teorico/Fudo.', fnt(9, it=True, color=MUTED), box=False)
put(ws, 47, 1, 'Descarte maximo tolerado en productos (%)', fnt(10, True))
put(ws, 47, 2, 0.03, FIN, IN_FILL, PCT, hor='center')
put(ws, 47, 3, 'Sobre el total elaborado de cada producto.', fnt(9, it=True, color=MUTED), box=False)
put(ws, 48, 1, 'Diferencia maxima tolerada sin explicar (%)', fnt(10, True))
put(ws, 48, 2, 0.02, FIN, IN_FILL, PCT, hor='center')
put(ws, 48, 3, 'Elaborado menos vendido, consumos y descartes.', fnt(9, it=True, color=MUTED), box=False)
ws.freeze_panes = 'A5'
imprimir(ws, 4, 48, landscape=False, repeat='1:3')
TOL_INS, TOL_DESC, TOL_DIF = 'Config!$B$46', 'Config!$B$47', 'Config!$B$48'

# =========================================================================
# Hoja INSUMOS (maestro) — 159 insumos de produccion
# A Area | B Insumo | C Categoria | D Proveedor | E Unidad | F Cont.x bulto
# G Critico | H Costo unit | I Clave
# =========================================================================
ws = SH['Insumos']
anchos(ws, [17, 34, 18, 15, 9, 12, 9, 13, 40])
titulo(ws, 9, 'MAESTRO DE INSUMOS DE PRODUCCION',
       'Traido del catalogo real de la app de Pedidos de Produccion (%d insumos, %d areas).' % (NINS, len(AREAS)),
       'REVISAR ANTES DE ARRANCAR: la unidad y el contenido por bulto fueron inferidos del nombre. Corrija lo que no coincida con el envase real.')
cabecera(ws, 4, ['AREA', 'INSUMO', 'CATEGORIA', 'PROVEEDOR', 'UNIDAD\nBASE',
                 'CONTENIDO\nPOR BULTO', 'CRITICO\n(conteo diario)', 'COSTO POR\nUNIDAD $', 'CLAVE (no tocar)'], 32)
for i, (ar, nm, cat, prov, un, bul, cri) in enumerate(INS):
    r = IR0 + i
    put(ws, r, 1, ar, fnt(9))
    put(ws, r, 2, nm, fnt(10, True))
    put(ws, r, 3, cat, FIN, IN_FILL, hor='center')
    put(ws, r, 4, prov, FIN, IN_FILL, hor='center')
    put(ws, r, 5, un,  FIN, IN_FILL, hor='center')
    put(ws, r, 6, bul, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 7, cri, FIN, IN_FILL, hor='center')
    put(ws, r, 8, None, FIN, IN_FILL, MON)
    put(ws, r, 9, '=$A%d&" | "&$B%d' % (r, r), fnt(8, color=MUTED))
put(ws, IR1 + 2, 2, 'El COSTO POR UNIDAD se completa desde las facturas de la semana (hoja "7 Facturas"). '
                    'Sin costo, el analisis sale en cantidades; con costo, sale ademas en pesos.',
    fnt(9, it=True, color=MUTED), box=False, wrap=True)
ws.merge_cells(start_row=IR1 + 2, start_column=2, end_row=IR1 + 2, end_column=8)
bandear(ws, IR0, IR1, 9)
dv = DV_UNI(); ws.add_data_validation(dv); dv.add('E%d:E%d' % (IR0, IR1))
dv = DV_SN();  ws.add_data_validation(dv); dv.add('G%d:G%d' % (IR0, IR1))
ws.column_dimensions['I'].hidden = True
ws.freeze_panes = 'C5'
ws.auto_filter.ref = 'A4:H%d' % IR1
imprimir(ws, 8, IR1)

CLAVE_RNG = "Insumos!$I$%d:$I$%d" % (IR0, IR1)
BULTO_RNG = "Insumos!$F$%d:$F$%d" % (IR0, IR1)
UNID_RNG  = "Insumos!$E$%d:$E$%d" % (IR0, IR1)
COSTO_RNG = "Insumos!$H$%d:$H$%d" % (IR0, IR1)

# =========================================================================
# Hoja PRODUCTOS (maestro) — 141 productos
# =========================================================================
ws = SH['Productos']
anchos(ws, [17, 58, 10, 18, 40])
titulo(ws, 5, 'MAESTRO DE PRODUCTOS ELABORADOS',
       'Traido del catalogo real de la app (%d productos).' % sum(len(v) for v in PROD.values()),
       'Cocina San Luis no tiene productos en el catalogo: se cargan a mano en su planilla de produccion.')
cabecera(ws, 4, ['AREA', 'PRODUCTO', 'UNIDAD', 'DONDE SE ELABORA', 'CLAVE (no tocar)'], 24)
pr = IR0
PRODIDX = {}
for a in AREAS:
    for p in PROD[a['key']]:
        put(ws, pr, 1, ALAB[a['key']], fnt(9))
        put(ws, pr, 2, p['name'], fnt(10))
        put(ws, pr, 3, p.get('unit', 'u'), fnt(10), hor='center')
        put(ws, pr, 4, DONDE[a['key']], fnt(9), hor='center')
        put(ws, pr, 5, '=$A%d&" | "&$B%d' % (pr, pr), fnt(8, color=MUTED))
        PRODIDX[(a['key'], p['name'])] = pr
        pr += 1
PR1 = pr - 1
bandear(ws, IR0, PR1, 5)
ws.column_dimensions['E'].hidden = True
ws.freeze_panes = 'B5'
ws.auto_filter.ref = 'A4:D%d' % PR1
imprimir(ws, 4, PR1, landscape=False)

# =========================================================================
# Helper: cabecera de dos niveles (dia arriba, sub-columnas abajo)
# =========================================================================
def dias_cabecera(ws, r, col0, sublabels, fijas):
    """fijas = lista de (titulo, ancho) para las columnas previas, merge vertical r:r+1."""
    for i, (t, _) in enumerate(fijas, start=1):
        ws.merge_cells(start_row=r, start_column=i, end_row=r + 1, end_column=i)
        put(ws, r, i, t, fnt(9, True, 'FFFFFF'), SUBFILL, hor='center', wrap=True)
        put(ws, r + 1, i, None, fnt(9, True, 'FFFFFF'), SUBFILL)
    n = len(sublabels)
    for d in range(7):
        c0 = col0 + d * n
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + n - 1)
        put(ws, r, c0, '=IF(%s="","%s",UPPER("%s")&" "&TEXT(%s,"dd/mm"))' % (CFG_D[d], DIAS[d], DIAS[d][:3], CFG_D[d]),
            fnt(9, True, PRIMDK), ACCFILL, hor='center')
        for k in range(1, n):
            put(ws, r, c0 + k, None, fill=ACCFILL)
        for k, sl in enumerate(sublabels):
            put(ws, r + 1, c0 + k, sl, fnt(8, True, 'FFFFFF'), SUBFILL, hor='center', wrap=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 24

# =========================================================================
# Hojas 1 INV INICIAL y 6 INV FINAL — todos los insumos, bultos + parcial
# A Area | B Insumo | C Unidad | D Cont/bulto | E Bultos | F Parcial | G TOTAL
# H Conto | I Observaciones          (filas alineadas 1:1 con el maestro)
# =========================================================================
def hoja_inventario(nom, tit, sub):
    ws = SH[nom]
    anchos(ws, [16, 33, 8, 11, 11, 12, 13, 13, 30])
    titulo(ws, 9, tit, sub,
           'COMO CONTAR: bultos cerrados + lo que hay suelto del abierto. '
           'Ejemplo: 3 bolsas de 25 kg + 8,5 kg sueltos  ->  Bultos = 3, Parcial = 8,5, TOTAL = 83,5 kg (lo calcula la planilla).')
    cabecera(ws, 4, ['AREA', 'INSUMO', 'UNID.', 'CONTENIDO\nPOR BULTO', 'BULTOS\nCERRADOS',
                     'PARCIAL\n(suelto)', 'TOTAL\n(calculado)', 'CONTO\n(iniciales)', 'OBSERVACIONES'], 34)
    for i in range(NINS):
        r = IR0 + i
        put(ws, r, 1, '=Insumos!$A%d' % r, fnt(9, color='008000'))
        put(ws, r, 2, '=Insumos!$B%d' % r, Font(name=F, size=10, bold=True, color='008000'))
        put(ws, r, 3, '=Insumos!$E%d' % r, FLINK, hor='center')
        put(ws, r, 4, '=Insumos!$F%d' % r, FLINK, None, NUM, hor='center')
        put(ws, r, 5, None, FIN, IN_FILL, NUM, hor='center')
        put(ws, r, 6, None, FIN, IN_FILL, NUM, hor='center')
        put(ws, r, 7, '=IFERROR($E%d*$D%d,0)+IFERROR($F%d,0)' % (r, r, r), fnt(10, True), TOTFILL, NUM, hor='center')
        put(ws, r, 8, None, FIN, IN_FILL, hor='center')
        put(ws, r, 9, None, FIN, IN_FILL)
    bandear(ws, IR0, IR1, 4)
    # marca en rojo claro los insumos que todavia no se contaron
    ws.conditional_formatting.add('G%d:G%d' % (IR0, IR1),
        FormulaRule(formula=['AND($E%d="",$F%d="")' % (IR0, IR0)],
                    fill=PatternFill('solid', fgColor=DANGBG), stopIfTrue=False))
    ws.freeze_panes = 'C5'
    ws.auto_filter.ref = 'A4:I%d' % IR1
    imprimir(ws, 9, IR1)
    return ws

hoja_inventario('1 Inv Inicial', 'PLANILLA 1 · INVENTARIO INICIAL DE INSUMOS',
    'Se cuenta el LUNES antes de arrancar a producir. Una hoja por area: filtre por AREA e imprima.')
hoja_inventario('6 Inv Final', 'PLANILLA 6 · INVENTARIO FINAL DE INSUMOS',
    'Se cuenta el DOMINGO al cierre, con el mismo criterio que el inicial. Misma persona si es posible.')

# =========================================================================
# Hoja 2 INGRESOS — mercaderia que entra durante la semana
# =========================================================================
ws = SH['2 Ingresos']
anchos(ws, [12, 16, 33, 15, 16, 10, 11, 11, 13, 14, 26])
titulo(ws, 11, 'PLANILLA 2 · INGRESOS DE MERCADERIA DE LA SEMANA',
       'Se anota CADA VEZ que entra mercaderia a un area, en el momento en que entra. Sin esto no se puede calcular el consumo.',
       'La fila 5 es un EJEMPLO: borrela antes de empezar. Cargue desde la fila 6.')
cabecera(ws, 4, ['FECHA', 'AREA', 'INSUMO', 'PROVEEDOR', 'N° REMITO /\nFACTURA', 'BULTOS',
                 'CONTENIDO\nPOR BULTO', 'PARCIAL', 'TOTAL\n(calculado)', 'RECIBIO\n(iniciales)',
                 'CONTROL AUTOMATICO'], 34)
ING0, ING1 = 6, 255
EJ = fnt(9, it=True, color=MUTED)
ejemplo = [None, 'Panaderia', 'Harina 000 x25kg', 'Molino XX', 'R-0001234', 4, None, 12.5, None, 'JP', None]
for c, v in enumerate(ejemplo, start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG), hor='center' if c in (6, 7, 8, 9) else None)
put(ws, 5, 1, '=IF(Config!$B$6="","",Config!$B$6)', EJ, PatternFill('solid', fgColor=INFOBG), DATE, hor='center')
put(ws, 5, 7, '=IFERROR(INDEX(%s,MATCH($B5&" | "&$C5,%s,0)),"")' % (BULTO_RNG, CLAVE_RNG), EJ,
    PatternFill('solid', fgColor=INFOBG), NUM, hor='center')
put(ws, 5, 9, '=IFERROR($F5*$G5,0)+IFERROR($H5,0)', EJ, PatternFill('solid', fgColor=INFOBG), NUM, hor='center')
put(ws, 5, 11, 'EJEMPLO - borrar esta fila', fnt(9, True, '1565C0'), PatternFill('solid', fgColor=INFOBG))
for r in range(ING0, ING1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL, DATE, hor='center')
    put(ws, r, 2, None, FIN, IN_FILL, hor='center')
    put(ws, r, 3, None, FIN, IN_FILL)
    put(ws, r, 4, None, FIN, IN_FILL, hor='center')
    put(ws, r, 5, None, FIN, IN_FILL, hor='center')
    put(ws, r, 6, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 7, '=IFERROR(INDEX(%s,MATCH($B%d&" | "&$C%d,%s,0)),"")' % (BULTO_RNG, r, r, CLAVE_RNG),
        FLINK, None, NUM, hor='center')
    put(ws, r, 8, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 9, '=IFERROR($F%d*$G%d,0)+IFERROR($H%d,0)' % (r, r, r), fnt(10, True), TOTFILL, NUM, hor='center')
    put(ws, r, 10, None, FIN, IN_FILL, hor='center')
    put(ws, r, 11, '=IF($C%d="","",IF(COUNTIF(%s,$B%d&" | "&$C%d)=0,"REVISAR: area+insumo no existe en el maestro","OK"))'
        % (r, CLAVE_RNG, r, r), fnt(9), hor='center')
ws.conditional_formatting.add('K%d:K%d' % (ING0, ING1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG)))
ws.conditional_formatting.add('K%d:K%d' % (ING0, ING1),
    FormulaRule(formula=['LEFT($K%d,7)="REVISAR"' % ING0],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
dv = DV_AREA(); ws.add_data_validation(dv); dv.add('B%d:B%d' % (ING0, ING1))
dv = DV_INS();  ws.add_data_validation(dv); dv.add('C%d:C%d' % (ING0, ING1))
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:K%d' % ING1
imprimir(ws, 11, ING1)
ING_TOT, ING_AR, ING_IN, ING_FE = ('%d' % ING0), None, None, None
R_ING_TOT = "'2 Ingresos'!$I$%d:$I$%d" % (ING0, ING1)
R_ING_AR  = "'2 Ingresos'!$B$%d:$B$%d" % (ING0, ING1)
R_ING_IN  = "'2 Ingresos'!$C$%d:$C$%d" % (ING0, ING1)
R_ING_FE  = "'2 Ingresos'!$A$%d:$A$%d" % (ING0, ING1)

def cols_finales(ws, r, cols):
    """cols = [(col_idx, titulo)] con merge vertical r:r+1."""
    for ci, t in cols:
        ws.merge_cells(start_row=r, start_column=ci, end_row=r + 1, end_column=ci)
        put(ws, r, ci, t, fnt(9, True, 'FFFFFF'), SUBFILL, hor='center', wrap=True)
        put(ws, r + 1, ci, None, fnt(9, True, 'FFFFFF'), SUBFILL)

# =========================================================================
# Hoja 3 CONTEO DIARIO — solo insumos CRITICOS, un bloque por dia
# =========================================================================
CRIT = [(IR0 + i, x[0], x[1]) for i, x in enumerate(INS) if x[6] == 'SI']
ws = SH['3 Conteo Diario']
anchos(ws, [16, 33, 8, 11] + [9, 9] * 7 + [11] * 7)
titulo(ws, 18, 'PLANILLA 3 · CONTEO DIARIO DE INSUMOS CRITICOS',
       'Cada jefe de area cuenta SOLO sus insumos criticos, al cierre de su jornada. %d insumos en total (unos 10 por area).' % len(CRIT),
       'Filtre por AREA e imprima una hoja por area. Mismo criterio que el inventario: bultos cerrados + suelto.')
dias_cabecera(ws, 4, 5, ['Bultos', 'Parcial'],
              [('AREA', 0), ('INSUMO', 0), ('UNID.', 0), ('CONTENIDO\nPOR BULTO', 0)])
cols_finales(ws, 4, [(19 + d, 'TOTAL\n' + DIAS[d][:3]) for d in range(7)])
CD0 = 6
for i, (mr, ar, nm) in enumerate(CRIT):
    r = CD0 + i
    put(ws, r, 1, '=Insumos!$A%d' % mr, fnt(9, color='008000'))
    put(ws, r, 2, '=Insumos!$B%d' % mr, Font(name=F, size=10, bold=True, color='008000'))
    put(ws, r, 3, '=Insumos!$E%d' % mr, FLINK, hor='center')
    put(ws, r, 4, '=Insumos!$F%d' % mr, FLINK, None, NUM, hor='center')
    for d in range(7):
        put(ws, r, 5 + 2 * d, None, FIN, IN_FILL, NUM, hor='center')
        put(ws, r, 6 + 2 * d, None, FIN, IN_FILL, NUM, hor='center')
        b, p = CL(5 + 2 * d), CL(6 + 2 * d)
        put(ws, r, 19 + d, '=IF(AND(%s%d="",%s%d=""),"",IFERROR(%s%d*$D%d,0)+IFERROR(%s%d,0))'
            % (b, r, p, r, b, r, r, p, r), fnt(9), None, NUM, hor='center')
CD1 = CD0 + len(CRIT) - 1
bandear(ws, CD0, CD1, 4)
ws.freeze_panes = 'C6'
ws.auto_filter.ref = 'A5:R%d' % CD1
imprimir(ws, 18, CD1, repeat='1:5')

# =========================================================================
# Hojas 4 PROD <AREA> — produccion diaria por area (elaborado + descarte)
# =========================================================================
PRODROW = {}   # (area_key, indice) -> fila
PRODLEN = {}
for a in AREAS:
    k = a['key']
    ws = SH[PRODSH[k]]
    anchos(ws, [52, 8] + [10, 9] * 7 + [12, 12, 10])
    nprod = len(PROD[k]); nblank = EXTRA[k]; total = nprod + nblank
    titulo(ws, 19, 'PLANILLA 4 · PRODUCCION DIARIA — %s' % a['label'].upper(),
           'Jefe de area: ____________________     Semana: ____/____ al ____/____     Se elabora en: %s' % DONDE[k],
           'ELABORADO = lo que salio bueno.  DESCARTE = lo que se quemo, salio mal, se cayo o se tiro. '
           'El detalle del motivo de cada descarte va en la planilla 5.')
    dias_cabecera(ws, 4, 3, ['ELABOR.', 'DESCAR.'], [('PRODUCTO', 0), ('UNID.', 0)])
    cols_finales(ws, 4, [(17, 'TOTAL\nELABORADO'), (18, 'TOTAL\nDESCARTE'), (19, '%\nDESC.')])
    PR = 6
    for i in range(total):
        r = PR + i
        if i < nprod:
            put(ws, r, 1, PROD[k][i]['name'], fnt(10))
            put(ws, r, 2, PROD[k][i].get('unit', 'u'), fnt(9), hor='center')
        else:
            put(ws, r, 1, None, FIN, IN_FILL)
            put(ws, r, 2, None, FIN, IN_FILL, hor='center')
        for d in range(7):
            put(ws, r, 3 + 2 * d, None, FIN, IN_FILL, NUM0, hor='center')
            put(ws, r, 4 + 2 * d, None, FIN, IN_FILL, NUM0, hor='center')
        put(ws, r, 17, '=SUM(%s)' % ','.join('%s%d' % (CL(3 + 2 * d), r) for d in range(7)),
            fnt(10, True), TOTFILL, NUM0, hor='center')
        put(ws, r, 18, '=SUM(%s)' % ','.join('%s%d' % (CL(4 + 2 * d), r) for d in range(7)),
            fnt(10, True), TOTFILL, NUM0, hor='center')
        put(ws, r, 19, '=IF($Q%d=0,"",$R%d/$Q%d)' % (r, r, r), fnt(10, True), TOTFILL, PCT, hor='center')
    LR = PR + total - 1
    PRODROW[k] = (PR, LR, nprod)
    PRODLEN[k] = total
    # totales de la planilla
    tr = LR + 1
    put(ws, tr, 1, 'TOTALES DEL AREA', fnt(10, True, 'FFFFFF'), SUBFILL)
    put(ws, tr, 2, None, fill=SUBFILL)
    for c in range(3, 20):
        put(ws, tr, c, '=SUM(%s%d:%s%d)' % (CL(c), PR, CL(c), LR) if c != 19 else
            '=IF($Q%d=0,"",$R%d/$Q%d)' % (tr, tr, tr),
            fnt(10, True, 'FFFFFF'), SUBFILL, PCT if c == 19 else NUM0, hor='center')
    ws.conditional_formatting.add('S%d:S%d' % (PR, LR),
        FormulaRule(formula=['AND($S%d<>"",$S%d>%s)' % (PR, PR, TOL_DESC)],
                    fill=PatternFill('solid', fgColor=DANGBG), font=fnt(10, True, 'C0392B')))
    bandear(ws, PR, LR, 2)
    ws.freeze_panes = 'C6'
    imprimir(ws, 19, tr, repeat='1:5')

# =========================================================================
# Hoja 5 DESCARTES — detalle con motivo
# =========================================================================
ws = SH['5 Descartes']
anchos(ws, [12, 17, 46, 11, 9, 25, 16, 30])
titulo(ws, 8, 'PLANILLA 5 · DETALLE DE DESCARTES, QUEMADO Y TIRADO',
       'Una linea por cada cosa que se descarta, con el motivo. Es la planilla que explica los faltantes: lo que no se anota aca aparece despues como perdida sin justificar.',
       'La fila 5 es un EJEMPLO: borrela antes de empezar. Cargue desde la fila 6.')
cabecera(ws, 4, ['FECHA', 'AREA', 'PRODUCTO O INSUMO', 'CANTIDAD', 'UNID.', 'MOTIVO',
                 'RESPONSABLE', 'OBSERVACIONES'], 30)
DES0, DES1 = 6, 205
ej = [None, 'Panaderia', 'Medialunas', 18, 'u', 'Quemado / mal cocido', 'JP', 'Horno con temperatura alta']
for c, v in enumerate(ej, start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG), hor='center' if c in (4, 5) else None)
put(ws, 5, 1, '=IF(Config!$B$6="","",Config!$B$6)', EJ, PatternFill('solid', fgColor=INFOBG), DATE, hor='center')
for r in range(DES0, DES1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL, DATE, hor='center')
    put(ws, r, 2, None, FIN, IN_FILL, hor='center')
    put(ws, r, 3, None, FIN, IN_FILL)
    put(ws, r, 4, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 5, None, FIN, IN_FILL, hor='center')
    put(ws, r, 6, None, FIN, IN_FILL)
    put(ws, r, 7, None, FIN, IN_FILL, hor='center')
    put(ws, r, 8, None, FIN, IN_FILL)
for dv, rng in ((DV_AREA(), 'B'), (DV_MOT(), 'F'), (DV_UNI(), 'E')):
    ws.add_data_validation(dv); dv.add('%s%d:%s%d' % (rng, DES0, rng, DES1))
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:H%d' % DES1
imprimir(ws, 8, DES1)
R_DES_CANT = "'5 Descartes'!$D$%d:$D$%d" % (DES0, DES1)
R_DES_AR   = "'5 Descartes'!$B$%d:$B$%d" % (DES0, DES1)
R_DES_IT   = "'5 Descartes'!$C$%d:$C$%d" % (DES0, DES1)
R_DES_MOT  = "'5 Descartes'!$F$%d:$F$%d" % (DES0, DES1)

# =========================================================================
# Hoja 7 FACTURAS DE COMPRA
# =========================================================================
ws = SH['7 Facturas']
anchos(ws, [12, 20, 16, 17, 33, 13, 8, 15, 15, 26])
titulo(ws, 10, 'CARGA 7 · FACTURAS DE COMPRA DE LA SEMANA',
       'Se carga DESPUES de la semana, con las facturas en la mano. La CANTIDAD va en la unidad base del insumo (kg, u, L), no en bultos.',
       'La fila 5 es un EJEMPLO: borrela antes de empezar. Cargue desde la fila 6.')
cabecera(ws, 4, ['FECHA', 'PROVEEDOR', 'N° FACTURA', 'AREA', 'INSUMO', 'CANTIDAD\n(unidad base)',
                 'UNID.', 'IMPORTE TOTAL $', 'COSTO UNITARIO $\n(calculado)', 'CONTROL AUTOMATICO'], 34)
FAC0, FAC1 = 6, 255
ej = [None, 'Molino XX', 'A-0001-00012345', 'Panaderia', 'Harina 000 x25kg', 100, 'kg', 85000, None, None]
for c, v in enumerate(ej, start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG),
        MON if c == 8 else (NUM if c == 6 else None), hor='center' if c in (6, 7, 8) else None)
put(ws, 5, 1, '=IF(Config!$B$6="","",Config!$B$6)', EJ, PatternFill('solid', fgColor=INFOBG), DATE, hor='center')
put(ws, 5, 9, '=IFERROR($H5/$F5,"")', EJ, PatternFill('solid', fgColor=INFOBG), MON, hor='center')
put(ws, 5, 10, 'EJEMPLO - borrar esta fila', fnt(9, True, '1565C0'), PatternFill('solid', fgColor=INFOBG))
for r in range(FAC0, FAC1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL, DATE, hor='center')
    put(ws, r, 2, None, FIN, IN_FILL)
    put(ws, r, 3, None, FIN, IN_FILL, hor='center')
    put(ws, r, 4, None, FIN, IN_FILL, hor='center')
    put(ws, r, 5, None, FIN, IN_FILL)
    put(ws, r, 6, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 7, '=IFERROR(INDEX(%s,MATCH($D%d&" | "&$E%d,%s,0)),"")' % (UNID_RNG, r, r, CLAVE_RNG),
        FLINK, None, None, hor='center')
    put(ws, r, 8, None, FIN, IN_FILL, MON, hor='center')
    put(ws, r, 9, '=IFERROR($H%d/$F%d,"")' % (r, r), fnt(10, True), TOTFILL, MON, hor='center')
    put(ws, r, 10, '=IF($E%d="","",IF(COUNTIF(%s,$D%d&" | "&$E%d)=0,"REVISAR: area+insumo no existe en el maestro","OK"))'
        % (r, CLAVE_RNG, r, r), fnt(9), hor='center')
ws.conditional_formatting.add('J%d:J%d' % (FAC0, FAC1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG)))
ws.conditional_formatting.add('J%d:J%d' % (FAC0, FAC1),
    FormulaRule(formula=['LEFT($J%d,7)="REVISAR"' % FAC0],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
for dv, col in ((DV_AREA(), 'D'), (DV_INS(), 'E')):
    ws.add_data_validation(dv); dv.add('%s%d:%s%d' % (col, FAC0, col, FAC1))
put(ws, FAC1 + 2, 5, 'TOTAL COMPRADO EN LA SEMANA', fnt(11, True, 'FFFFFF'), SUBFILL, hor='right')
put(ws, FAC1 + 2, 8, '=SUM($H$%d:$H$%d)' % (FAC0, FAC1), fnt(11, True, 'FFFFFF'), SUBFILL, MON, hor='center')
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:J%d' % FAC1
imprimir(ws, 10, FAC1)
R_FAC_CANT = "'7 Facturas'!$F$%d:$F$%d" % (FAC0, FAC1)
R_FAC_AR   = "'7 Facturas'!$D$%d:$D$%d" % (FAC0, FAC1)
R_FAC_IN   = "'7 Facturas'!$E$%d:$E$%d" % (FAC0, FAC1)
R_FAC_IMP  = "'7 Facturas'!$H$%d:$H$%d" % (FAC0, FAC1)

# =========================================================================
# Hoja 8 FUDO VENTAS
# =========================================================================
ws = SH['8 Fudo Ventas']
anchos(ws, [42, 14, 17, 46, 26])
titulo(ws, 5, 'CARGA 8a · VENTAS POR PRODUCTO SEGUN FUDO',
       'Pegue el export de Fudo en las columnas A y B. Despues indique a que AREA y a que PRODUCTO de Candela corresponde cada linea.',
       'Si un producto de Fudo no se mapea, no entra en la conciliacion. La fila 5 es un EJEMPLO: borrela.')
cabecera(ws, 4, ['PRODUCTO SEGUN FUDO\n(pegar del export)', 'UNIDADES\nVENDIDAS',
                 'AREA DE CANDELA', 'PRODUCTO DE CANDELA', 'CONTROL AUTOMATICO'], 34)
FV0, FV1 = 6, 305
for c, v in enumerate(['Medialuna de manteca', 640, 'Facturería', 'Medialunas', None], start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG), NUM0 if c == 2 else None,
        hor='center' if c in (2, 3) else None)
put(ws, 5, 5, 'EJEMPLO - borrar esta fila', fnt(9, True, '1565C0'), PatternFill('solid', fgColor=INFOBG))
for r in range(FV0, FV1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL)
    put(ws, r, 2, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 3, None, FIN, IN_FILL, hor='center')
    put(ws, r, 4, None, FIN, IN_FILL)
    put(ws, r, 5, '=IF($A%d="","",IF($D%d="","FALTA MAPEAR","OK"))' % (r, r), fnt(9), hor='center')
ws.conditional_formatting.add('E%d:E%d' % (FV0, FV1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG)))
ws.conditional_formatting.add('E%d:E%d' % (FV0, FV1),
    CellIsRule(operator='equal', formula=['"FALTA MAPEAR"'],
               fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
for dv, col in ((DV_AREA(), 'C'), (DV_PROD(), 'D')):
    ws.add_data_validation(dv); dv.add('%s%d:%s%d' % (col, FV0, col, FV1))
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:E%d' % FV1
imprimir(ws, 5, FV1, landscape=False)
R_FV_CANT = "'8 Fudo Ventas'!$B$%d:$B$%d" % (FV0, FV1)
R_FV_AR   = "'8 Fudo Ventas'!$C$%d:$C$%d" % (FV0, FV1)
R_FV_PR   = "'8 Fudo Ventas'!$D$%d:$D$%d" % (FV0, FV1)

# =========================================================================
# Hoja 8 FUDO STOCK
# =========================================================================
ws = SH['8 Fudo Stock']
anchos(ws, [38, 12, 12, 13, 12, 17, 33, 16, 40])
titulo(ws, 9, 'CARGA 8b · MOVIMIENTOS DE STOCK DE INSUMOS SEGUN FUDO',
       'Pegue el reporte de existencias/movimientos de Fudo en A:E. Despues mapee AREA e INSUMO de Candela.',
       'Si Fudo no informa egresos, la planilla los calcula: inicial + ingresos - final. La fila 5 es un EJEMPLO: borrela.')
cabecera(ws, 4, ['INSUMO SEGUN FUDO\n(pegar del export)', 'STOCK\nINICIAL', 'INGRESOS', 'EGRESOS',
                 'STOCK\nFINAL', 'AREA DE CANDELA', 'INSUMO DE CANDELA',
                 'CONSUMO S/FUDO\n(calculado)', 'CLAVE (no tocar)'], 34)
FS0, FS1 = 6, 205
for c, v in enumerate(['Harina 000', 120, 100, None, 95, 'Panadería', 'Harina 000 x25kg'], start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG), NUM if 2 <= c <= 5 else None,
        hor='center' if 2 <= c <= 6 else None)
put(ws, 5, 8, '=IF($D5<>"",$D5,IF(COUNT($B5,$E5)=2,$B5+IFERROR($C5,0)-$E5,""))', EJ,
    PatternFill('solid', fgColor=INFOBG), NUM, hor='center')
for r in range(FS0, FS1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL)
    for c in (2, 3, 4, 5):
        put(ws, r, c, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 6, None, FIN, IN_FILL, hor='center')
    put(ws, r, 7, None, FIN, IN_FILL)
    put(ws, r, 8, '=IF($D%d<>"",$D%d,IF(COUNT($B%d,$E%d)=2,$B%d+IFERROR($C%d,0)-$E%d,""))'
        % (r, r, r, r, r, r, r), fnt(10, True), TOTFILL, NUM, hor='center')
    put(ws, r, 9, '=IF($G%d="","",$F%d&" | "&$G%d)' % (r, r, r), fnt(8, color=MUTED))
for dv, col in ((DV_AREA(), 'F'), (DV_INS(), 'G')):
    ws.add_data_validation(dv); dv.add('%s%d:%s%d' % (col, FS0, col, FS1))
ws.column_dimensions['I'].hidden = True
ws.freeze_panes = 'B5'
ws.auto_filter.ref = 'A4:H%d' % FS1
imprimir(ws, 8, FS1)
R_FS_CONS  = "'8 Fudo Stock'!$H$%d:$H$%d" % (FS0, FS1)
R_FS_CLAVE = "'8 Fudo Stock'!$I$%d:$I$%d" % (FS0, FS1)

# =========================================================================
# Hoja 8 FUDO CONSUMOS
# =========================================================================
ws = SH['8 Fudo Consumos']
anchos(ws, [12, 46, 17, 12, 9, 24, 18, 28])
titulo(ws, 8, 'CARGA 8c · CONSUMOS INTERNOS Y CORTESIAS SEGUN FUDO',
       'Todo lo que se consumio sin venderse y quedo registrado en Fudo: consumo del personal, cortesias, pruebas.',
       'Esta hoja es la que separa el consumo legitimo del faltante real. La fila 5 es un EJEMPLO: borrela.')
cabecera(ws, 4, ['FECHA', 'PRODUCTO DE CANDELA', 'AREA', 'CANTIDAD', 'UNID.', 'TIPO',
                 'RESPONSABLE / SECTOR', 'OBSERVACIONES'], 30)
FC0, FC1 = 6, 255
for c, v in enumerate([None, 'Medialunas', 'Facturería', 24, 'u', 'Consumo del personal', 'Salon manana'], start=1):
    put(ws, 5, c, v, EJ, PatternFill('solid', fgColor=INFOBG), NUM if c == 4 else None,
        hor='center' if c in (3, 4, 5) else None)
put(ws, 5, 1, '=IF(Config!$B$6="","",Config!$B$6)', EJ, PatternFill('solid', fgColor=INFOBG), DATE, hor='center')
for r in range(FC0, FC1 + 1):
    put(ws, r, 1, None, FIN, IN_FILL, DATE, hor='center')
    put(ws, r, 2, None, FIN, IN_FILL)
    put(ws, r, 3, None, FIN, IN_FILL, hor='center')
    put(ws, r, 4, None, FIN, IN_FILL, NUM, hor='center')
    put(ws, r, 5, None, FIN, IN_FILL, hor='center')
    put(ws, r, 6, None, FIN, IN_FILL)
    put(ws, r, 7, None, FIN, IN_FILL, hor='center')
    put(ws, r, 8, None, FIN, IN_FILL)
for dv, col in ((DV_AREA(), 'C'), (DV_PROD(), 'B'), (DV_TIPO(), 'F'), (DV_UNI(), 'E')):
    ws.add_data_validation(dv); dv.add('%s%d:%s%d' % (col, FC0, col, FC1))
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:H%d' % FC1
imprimir(ws, 8, FC1)
R_FC_CANT = "'8 Fudo Consumos'!$D$%d:$D$%d" % (FC0, FC1)
R_FC_PR   = "'8 Fudo Consumos'!$B$%d:$B$%d" % (FC0, FC1)
R_FC_AR   = "'8 Fudo Consumos'!$C$%d:$C$%d" % (FC0, FC1)

# =========================================================================
# Hoja A · CONSUMO DE INSUMOS  (inicial + ingresos - final = consumo real)
# =========================================================================
ws = SH['A Consumo Insumos']
anchos(ws, [16, 33, 7, 11, 11, 11, 13, 12, 12, 12, 12, 10, 11, 14, 14, 34, 8, 8])
titulo(ws, 16, 'ANALISIS A · CONSUMO REAL DE INSUMOS DE LA SEMANA',
       'CONSUMO REAL = Stock inicial + Ingresos de la semana - Stock final.  Se compara contra las facturas y contra Fudo.',
       'Un CONSUMO NEGATIVO significa que entro mercaderia que nadie anoto. Un DESVIO ALTO contra Fudo es el candidato a faltante o a mal uso.')
cabecera(ws, 4, ['AREA', 'INSUMO', 'UNID.', 'STOCK\nINICIAL', 'INGRESOS\nS/PLANILLA',
                 'STOCK\nFINAL', 'CONSUMO\nREAL', 'COMPRAS\nS/FACTURAS', 'DIF. RECEP.\nVS FACTURA',
                 'CONSUMO\nS/FUDO', 'DIF. REAL\nVS FUDO', '% DESVIO',
                 'COSTO\nUNIT. $', 'CONSUMO\nVALORIZADO $', 'DESVIO\nVALORIZADO $', 'ALERTA',
                 'k%', 'k$'], 36)
for i in range(NINS):
    r = IR0 + i
    put(ws, r, 1, '=Insumos!$A%d' % r, fnt(9, color='008000'))
    put(ws, r, 2, '=Insumos!$B%d' % r, Font(name=F, size=10, bold=True, color='008000'))
    put(ws, r, 3, '=Insumos!$E%d' % r, FLINK, hor='center')
    put(ws, r, 4, "='1 Inv Inicial'!$G%d" % r, FLINK, None, NUM, hor='center')
    put(ws, r, 5, '=SUMIFS(%s,%s,$A%d,%s,$B%d)' % (R_ING_TOT, R_ING_AR, r, R_ING_IN, r),
        FCALC, None, NUM, hor='center')
    put(ws, r, 6, "='6 Inv Final'!$G%d" % r, FLINK, None, NUM, hor='center')
    put(ws, r, 7, '=$D%d+$E%d-$F%d' % (r, r, r), fnt(10, True), ACCFILL, NUM, hor='center')
    put(ws, r, 8, '=SUMIFS(%s,%s,$A%d,%s,$B%d)' % (R_FAC_CANT, R_FAC_AR, r, R_FAC_IN, r),
        FCALC, None, NUM, hor='center')
    put(ws, r, 9, '=$E%d-$H%d' % (r, r), FCALC, None, NUM, hor='center')
    put(ws, r, 10, '=IFERROR(INDEX(%s,MATCH($A%d&" | "&$B%d,%s,0)),"")' % (R_FS_CONS, r, r, R_FS_CLAVE),
        FCALC, None, NUM, hor='center')
    put(ws, r, 11, '=IF($J%d="","",$G%d-$J%d)' % (r, r, r), FCALC, None, NUM, hor='center')
    put(ws, r, 12, '=IF(OR($J%d="",$J%d=0),"",$K%d/$J%d)' % (r, r, r, r), fnt(10, True), None, PCT, hor='center')
    put(ws, r, 13, '=Insumos!$H%d' % r, FLINK, None, MON, hor='center')
    put(ws, r, 14, '=IF($M%d="","",$G%d*$M%d)' % (r, r, r), FCALC, None, MON, hor='center')
    put(ws, r, 15, '=IF(OR($M%d="",$K%d=""),"",$K%d*$M%d)' % (r, r, r, r), fnt(10, True), None, MON, hor='center')
    put(ws, r, 16, ('=IF(AND(\'1 Inv Inicial\'!$E{r}="",\'1 Inv Inicial\'!$F{r}=""),"Falta inventario inicial",'
                    'IF(AND(\'6 Inv Final\'!$E{r}="",\'6 Inv Final\'!$F{r}=""),"Falta inventario final",'
                    'IF($G{r}<0,"Consumo negativo: entro mercaderia sin registrar",'
                    'IF(AND($L{r}<>"",ABS($L{r})>{tol}),"Desvio alto vs Fudo: investigar",'
                    'IF(AND($H{r}>0,ABS($I{r})/$H{r}>{tol}),"Lo recibido no coincide con la factura",'
                    'IF(AND($J{r}="",$G{r}>0),"Sin dato de Fudo para comparar","OK"))))))'
                   ).format(r=r, tol=TOL_INS), fnt(9), None, None, wrap=True)
    put(ws, r, 17, '=IF($L%d="","",ABS($L%d)+ROW()/1000000)' % (r, r), fnt(8, color=MUTED), None, NUM)
    put(ws, r, 18, '=IF($O%d="","",ABS($O%d)+ROW()/1000000)' % (r, r), fnt(8, color=MUTED), None, NUM)
bandear(ws, IR0, IR1, 3)
TR = IR1 + 1
put(ws, TR, 1, 'TOTALES', fnt(11, True, 'FFFFFF'), SUBFILL)
for c in (2, 3):
    put(ws, TR, c, None, fill=SUBFILL)
for c in (14, 15):
    put(ws, TR, c, '=SUM(%s$%d:%s$%d)' % (CL(c), IR0, CL(c), IR1), fnt(11, True, 'FFFFFF'), SUBFILL, MON, hor='center')
for c in range(4, 14):
    put(ws, TR, c, None, fill=SUBFILL)
put(ws, TR, 16, '=COUNTIF($P$%d:$P$%d,"OK")&" de %d insumos OK"' % (IR0, IR1, NINS),
    fnt(10, True, 'FFFFFF'), SUBFILL, hor='center')
ws.conditional_formatting.add('P%d:P%d' % (IR0, IR1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG), font=fnt(9, color='2E7D45')))
ws.conditional_formatting.add('P%d:P%d' % (IR0, IR1),
    FormulaRule(formula=['AND($P%d<>"",$P%d<>"OK")' % (IR0, IR0)],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
ws.conditional_formatting.add('L%d:L%d' % (IR0, IR1),
    FormulaRule(formula=['AND($L%d<>"",ABS($L%d)>%s)' % (IR0, IR0, TOL_INS)],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(10, True, 'C0392B')))
ws.column_dimensions['Q'].hidden = True
ws.column_dimensions['R'].hidden = True
ws.freeze_panes = 'C5'
ws.auto_filter.ref = 'A4:P%d' % IR1
imprimir(ws, 16, TR)
R_AC_INS = "'A Consumo Insumos'!$B$%d:$B$%d" % (IR0, IR1)
R_AC_AR  = "'A Consumo Insumos'!$A$%d:$A$%d" % (IR0, IR1)
R_AC_K1  = "'A Consumo Insumos'!$Q$%d:$Q$%d" % (IR0, IR1)
R_AC_K2  = "'A Consumo Insumos'!$R$%d:$R$%d" % (IR0, IR1)
R_AC_PCT = "'A Consumo Insumos'!$L$%d:$L$%d" % (IR0, IR1)
R_AC_DIF = "'A Consumo Insumos'!$K$%d:$K$%d" % (IR0, IR1)
R_AC_DV  = "'A Consumo Insumos'!$O$%d:$O$%d" % (IR0, IR1)
R_AC_UNI = "'A Consumo Insumos'!$C$%d:$C$%d" % (IR0, IR1)
R_AC_ALE = "'A Consumo Insumos'!$P$%d:$P$%d" % (IR0, IR1)

# =========================================================================
# Hoja B · CONCILIACION DE PRODUCTOS
# =========================================================================
ws = SH['B Concil Productos']
anchos(ws, [16, 50, 7, 12, 12, 13, 12, 13, 13, 15, 10, 34, 8])
titulo(ws, 12, 'ANALISIS B · CONCILIACION DE PRODUCTOS ELABORADOS',
       'ELABORADO - VENDIDO (Fudo) - CONSUMO INTERNO (Fudo) - DESCARTE DECLARADO = DIFERENCIA SIN EXPLICAR.',
       'OJO: no se lleva stock de producto terminado, asi que la diferencia incluye lo que quedo en exhibicion. Lea la TENDENCIA y los casos grandes, no el numero exacto.')
cabecera(ws, 4, ['AREA', 'PRODUCTO', 'UNID.', 'ELABORADO\n(planilla 4)', 'VENDIDO\n(Fudo)',
                 'CONSUMO INTERNO\n(Fudo)', 'DESCARTE\n(planilla 4)', 'DESCARTE DETALLADO\n(planilla 5)',
                 'DESCARTE SIN\nDETALLAR', 'DIFERENCIA SIN\nEXPLICAR', '% S/ELAB.', 'ALERTA', 'k'], 38)
BR = 6
BCROW = []
for a in AREAS:
    k = a['key']; sn = PRODSH[k]; pr0, pr1, npr = PRODROW[k]
    for j in range(pr1 - pr0 + 1):
        r = BR + len(BCROW)
        src = pr0 + j
        put(ws, r, 1, ALAB[k], fnt(9))
        put(ws, r, 2, "='%s'!$A%d" % (sn, src), Font(name=F, size=10, color='008000'))
        put(ws, r, 3, "='%s'!$B%d" % (sn, src), FLINK, hor='center')
        put(ws, r, 4, "='%s'!$Q%d" % (sn, src), FLINK, None, NUM0, hor='center')
        put(ws, r, 5, '=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$A%d))' % (r, R_FV_CANT, R_FV_PR, r, R_FV_AR, r),
            FCALC, None, NUM0, hor='center')
        put(ws, r, 6, '=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$A%d))' % (r, R_FC_CANT, R_FC_PR, r, R_FC_AR, r),
            FCALC, None, NUM0, hor='center')
        put(ws, r, 7, "='%s'!$R%d" % (sn, src), FLINK, None, NUM0, hor='center')
        put(ws, r, 8, '=IF($B%d="","",SUMIFS(%s,%s,$B%d,%s,$A%d))' % (r, R_DES_CANT, R_DES_IT, r, R_DES_AR, r),
            FCALC, None, NUM0, hor='center')
        put(ws, r, 9, '=IF($B%d="","",$G%d-$H%d)' % (r, r, r), FCALC, None, NUM0, hor='center')
        put(ws, r, 10, '=IF($B%d="","",$D%d-$E%d-$F%d-$G%d)' % (r, r, r, r, r), fnt(10, True), ACCFILL, NUM0, hor='center')
        put(ws, r, 11, '=IF(OR($B%d="",$D%d=0),"",$J%d/$D%d)' % (r, r, r, r), fnt(10, True), None, PCT, hor='center')
        put(ws, r, 12, ('=IF($B{r}="","",IF($D{r}=0,"Sin produccion registrada",'
                        'IF($E{r}=0,"Sin ventas mapeadas en Fudo",'
                        'IF(AND($K{r}<>"",ABS($K{r})>{tol}),IF($J{r}>0,'
                        '"Salida sin explicar: consumo no registrado o faltante",'
                        '"Se vendio mas de lo elaborado: revisar stock previo"),'
                        'IF($I{r}<>0,"Hay descarte sin detallar en la planilla 5","OK")))))'
                       ).format(r=r, tol=TOL_DIF), fnt(9), None, None, wrap=True)
        put(ws, r, 13, '=IF($K%d="","",ABS($K%d)+ROW()/1000000)' % (r, r), fnt(8, color=MUTED), None, NUM)
        BCROW.append(r)
BR1 = BR + len(BCROW) - 1
bandear(ws, BR, BR1, 3)
TR2 = BR1 + 1
put(ws, TR2, 1, 'TOTALES', fnt(11, True, 'FFFFFF'), SUBFILL)
for c in (2, 3, 11, 13):
    put(ws, TR2, c, None, fill=SUBFILL)
for c in (4, 5, 6, 7, 8, 9, 10):
    put(ws, TR2, c, '=SUM(%s$%d:%s$%d)' % (CL(c), BR, CL(c), BR1), fnt(11, True, 'FFFFFF'), SUBFILL, NUM0, hor='center')
put(ws, TR2, 11, '=IF($D%d=0,"",$J%d/$D%d)' % (TR2, TR2, TR2), fnt(11, True, 'FFFFFF'), SUBFILL, PCT, hor='center')
put(ws, TR2, 12, '=COUNTIF($L$%d:$L$%d,"OK")&" productos OK"' % (BR, BR1), fnt(10, True, 'FFFFFF'), SUBFILL, hor='center')
ws.conditional_formatting.add('L%d:L%d' % (BR, BR1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG), font=fnt(9, color='2E7D45')))
ws.conditional_formatting.add('L%d:L%d' % (BR, BR1),
    FormulaRule(formula=['AND($L%d<>"",$L%d<>"OK")' % (BR, BR)],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
ws.column_dimensions['M'].hidden = True
ws.freeze_panes = 'C5'
ws.auto_filter.ref = 'A4:L%d' % BR1
imprimir(ws, 12, TR2)
R_BC_PR  = "'B Concil Productos'!$B$%d:$B$%d" % (BR, BR1)
R_BC_AR  = "'B Concil Productos'!$A$%d:$A$%d" % (BR, BR1)
R_BC_K   = "'B Concil Productos'!$M$%d:$M$%d" % (BR, BR1)
R_BC_DIF = "'B Concil Productos'!$J$%d:$J$%d" % (BR, BR1)
R_BC_PCT = "'B Concil Productos'!$K$%d:$K$%d" % (BR, BR1)
R_BC_ELA = "'B Concil Productos'!$D$%d:$D$%d" % (BR, BR1)
R_BC_DES = "'B Concil Productos'!$G$%d:$G$%d" % (BR, BR1)
R_BC_ALE = "'B Concil Productos'!$L$%d:$L$%d" % (BR, BR1)

# =========================================================================
# Hoja C · CONSUMO DIARIO DE CRITICOS (para ubicar QUE DIA se produce el desvio)
# =========================================================================
ws = SH['C Consumo Diario']
anchos(ws, [16, 33, 7] + [11] * 7 + [12, 11, 13, 34])
titulo(ws, 14, 'ANALISIS C · CONSUMO DIA POR DIA DE LOS INSUMOS CRITICOS',
       'Consumo del dia = stock del dia anterior + ingresos de ese dia - stock al cierre del dia. Sale de la planilla 3.',
       'Sirve para ubicar EL DIA exacto del desvio. Un consumo negativo = entro mercaderia que nadie anoto ese dia.')
hdr = ['AREA', 'INSUMO', 'UNID.'] + [None] * 7 + ['TOTAL\nSEMANA', 'PROMEDIO\nPOR DIA', 'DIA DE MAYOR\nCONSUMO', 'ALERTA']
cabecera(ws, 4, hdr, 34)
for d in range(7):
    put(ws, 4, 4 + d, '=IF(%s="","%s",UPPER("%s")&" "&TEXT(%s,"dd/mm"))' % (CFG_D[d], DIAS[d][:3], DIAS[d][:3], CFG_D[d]),
        fnt(9, True, 'FFFFFF'), SUBFILL, hor='center', wrap=True)
CC0 = 5
for i, (mr, ar, nm) in enumerate(CRIT):
    r = CC0 + i
    cd = CD0 + i
    put(ws, r, 1, '=Insumos!$A%d' % mr, fnt(9, color='008000'))
    put(ws, r, 2, '=Insumos!$B%d' % mr, Font(name=F, size=10, bold=True, color='008000'))
    put(ws, r, 3, '=Insumos!$E%d' % mr, FLINK, hor='center')
    for d in range(7):
        now = "'3 Conteo Diario'!$%s%d" % (CL(19 + d), cd)
        ing = '=SUMIFS(%s,%s,$A%d,%s,$B%d,%s,%s)' % (R_ING_TOT, R_ING_AR, r, R_ING_IN, r, R_ING_FE, CFG_D[d])
        ing = ing[1:]
        if d == 0:
            prev = "'1 Inv Inicial'!$G%d" % mr
            f = '=IF(%s="","",%s+%s-%s)' % (now, prev, ing, now)
        else:
            prev = "'3 Conteo Diario'!$%s%d" % (CL(18 + d), cd)
            f = '=IF(OR(%s="",%s=""),"",%s+%s-%s)' % (prev, now, prev, ing, now)
        put(ws, r, 4 + d, f, fnt(9), None, NUM, hor='center')
    put(ws, r, 11, '=IF(COUNT($D%d:$J%d)=0,"",SUM($D%d:$J%d))' % (r, r, r, r), fnt(10, True), ACCFILL, NUM, hor='center')
    put(ws, r, 12, '=IF(COUNT($D%d:$J%d)=0,"",AVERAGE($D%d:$J%d))' % (r, r, r, r), FCALC, None, NUM, hor='center')
    put(ws, r, 13, '=IF(COUNT($D%d:$J%d)=0,"",IFERROR(INDEX($D$4:$J$4,MATCH(MAX($D%d:$J%d),$D%d:$J%d,0)),""))'
        % (r, r, r, r, r, r), fnt(9), None, None, hor='center')
    put(ws, r, 14, ('=IF(COUNT($D{r}:$J{r})=0,"Sin conteos cargados",'
                    'IF(MIN($D{r}:$J{r})<0,"Consumo negativo: ingreso sin registrar",'
                    'IF(AND($L{r}>0,MAX($D{r}:$J{r})>2*$L{r}),"Pico atipico el "&$M{r},"OK")))').format(r=r),
        fnt(9), None, None, wrap=True)
CC1 = CC0 + len(CRIT) - 1
bandear(ws, CC0, CC1, 3)
ws.conditional_formatting.add('D%d:J%d' % (CC0, CC1),
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
ws.conditional_formatting.add('N%d:N%d' % (CC0, CC1),
    CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill('solid', fgColor=OKBG), font=fnt(9, color='2E7D45')))
ws.conditional_formatting.add('N%d:N%d' % (CC0, CC1),
    FormulaRule(formula=['AND($N%d<>"",$N%d<>"OK")' % (CC0, CC0)],
                fill=PatternFill('solid', fgColor=DANGBG), font=fnt(9, True, 'C0392B')))
ws.freeze_panes = 'C5'
ws.auto_filter.ref = 'A4:N%d' % CC1
imprimir(ws, 14, CC1)

# =========================================================================
# Hoja TABLERO
# =========================================================================
ws = SH['Tablero']
anchos(ws, [4, 34, 34, 15, 15, 15, 15, 10, 3])
titulo(ws, 8, 'TABLERO DE RESULTADOS DE LA SEMANA',
       'Se completa solo a medida que se cargan las planillas. Es la hoja para la reunion de cierre.',
       'Los limites de tolerancia se configuran en la hoja Config (seccion 5).')

def bloque(r, txt):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    c = ws.cell(row=r, column=2, value=txt)
    c.font = fnt(11, True, 'FFFFFF'); c.fill = SUBFILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r].height = 22

bloque(5, '1 · ESTADO DE CARGA  (antes de leer nada, esto tiene que estar completo)')
CARGA = [
    ('Insumos con inventario inicial contado', '=COUNTIF(\'1 Inv Inicial\'!$E$%d:$E$%d,">=0")+COUNTIF(\'1 Inv Inicial\'!$F$%d:$F$%d,">=0")-SUMPRODUCT((\'1 Inv Inicial\'!$E$%d:$E$%d<>"")*(\'1 Inv Inicial\'!$F$%d:$F$%d<>""))' % ((IR0, IR1) * 4), '=%d' % NINS),
    ('Insumos con inventario final contado', '=COUNTIF(\'6 Inv Final\'!$E$%d:$E$%d,">=0")+COUNTIF(\'6 Inv Final\'!$F$%d:$F$%d,">=0")-SUMPRODUCT((\'6 Inv Final\'!$E$%d:$E$%d<>"")*(\'6 Inv Final\'!$F$%d:$F$%d<>""))' % ((IR0, IR1) * 4), '=%d' % NINS),
    ('Lineas de ingreso de mercaderia cargadas', '=COUNTIF(%s,"<>")' % R_ING_AR, None),
    ('Lineas de factura cargadas', '=COUNTIF(%s,"<>")' % R_FAC_AR, None),
    ('Lineas de descarte cargadas', '=COUNTIF(%s,"<>")' % R_DES_AR, None),
    ('Lineas de venta de Fudo mapeadas', '=COUNTIF(%s,"<>")' % R_FV_PR, None),
    ('Lineas de venta de Fudo SIN mapear', '=COUNTIF(\'8 Fudo Ventas\'!$E$%d:$E$%d,"FALTA MAPEAR")' % (FV0, FV1), '=0'),
    ('Insumos con costo cargado (para valorizar)', '=COUNT(%s)' % COSTO_RNG, '=%d' % NINS),
]
cabecera(ws, 6, ['', 'CONCEPTO', '', 'CARGADO', 'ESPERADO', '', '', ''], 20)
for i, (lab, f, esp) in enumerate(CARGA):
    r = 7 + i
    put(ws, r, 2, lab, fnt(10), BANDFILL if i % 2 else None)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    put(ws, r, 3, None, fill=BANDFILL if i % 2 else None)
    put(ws, r, 4, f, fnt(10, True), None, NUM0, hor='center')
    put(ws, r, 5, esp, fnt(10, color=MUTED), None, NUM0, hor='center')
    put(ws, r, 6, None, box=False); put(ws, r, 7, None, box=False); put(ws, r, 8, None, box=False)

bloque(16, '2 · RESUMEN DE LA SEMANA')
KPI = [
    ('Consumo de insumos valorizado', "='A Consumo Insumos'!$N$%d" % TR, MON),
    ('Total comprado segun facturas', "='7 Facturas'!$H$%d" % (FAC1 + 2), MON),
    ('Desvio valorizado vs Fudo (lo que no cierra)', "='A Consumo Insumos'!$O$%d" % TR, MON),
    ('Insumos con alerta', '=COUNTIFS(%s,"<>OK",%s,"<>")' % (R_AC_ALE, R_AC_ALE), NUM0),
    ('Insumos con consumo negativo (ingreso sin registrar)', '=COUNTIF(%s,"Consumo negativo*")' % R_AC_ALE, NUM0),
    ('Total elaborado (unidades)', "='B Concil Productos'!$D$%d" % TR2, NUM0),
    ('Total descartado declarado (unidades)', "='B Concil Productos'!$G$%d" % TR2, NUM0),
    ('% de descarte sobre lo elaborado', "=IF('B Concil Productos'!$D$%d=0,\"\",'B Concil Productos'!$G$%d/'B Concil Productos'!$D$%d)" % (TR2, TR2, TR2), PCT),
    ('Consumo interno registrado en Fudo (unidades)', "='B Concil Productos'!$F$%d" % TR2, NUM0),
    ('Diferencia total sin explicar (unidades)', "='B Concil Productos'!$J$%d" % TR2, NUM0),
    ('Productos con alerta', '=COUNTIFS(%s,"<>OK",%s,"<>")' % (R_BC_ALE, R_BC_ALE), NUM0),
]
for i, (lab, f, fm) in enumerate(KPI):
    r = 17 + i
    put(ws, r, 2, lab, fnt(10), BANDFILL if i % 2 else None)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    put(ws, r, 3, None, fill=BANDFILL if i % 2 else None)
    put(ws, r, 4, f, fnt(11, True), ACCFILL, fm, hor='center')
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    put(ws, r, 5, None, fill=ACCFILL)
    for c in (6, 7, 8): put(ws, r, c, None, box=False)

bloque(30, '3 · LOS 12 INSUMOS CON MAYOR DESVIO  (candidatos a faltante o mal uso)')
cabecera(ws, 31, ['#', 'AREA', 'INSUMO', 'CONSUMO REAL', 'DIF. VS FUDO', '% DESVIO', 'DESVIO $', 'k'], 22)
for n in range(1, 13):
    r = 31 + n
    put(ws, r, 1, n, fnt(9, color=MUTED), BANDFILL if n % 2 else None, hor='center')
    put(ws, r, 8, '=IFERROR(LARGE(%s,%d),"")' % (R_AC_K1, n), fnt(8, color=MUTED), None, NUM)
    m = 'MATCH($H%d,%s,0)' % (r, R_AC_K1)
    put(ws, r, 2, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_AC_AR, m), fnt(9), BANDFILL if n % 2 else None)
    put(ws, r, 3, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_AC_INS, m), fnt(10, True), BANDFILL if n % 2 else None)
    put(ws, r, 4, '=IF($H%d="","",INDEX(%s,%s))' % (r, "'A Consumo Insumos'!$G$%d:$G$%d" % (IR0, IR1), m),
        FCALC, None, NUM, hor='center')
    put(ws, r, 5, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_AC_DIF, m), FCALC, None, NUM, hor='center')
    put(ws, r, 6, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_AC_PCT, m), fnt(10, True, 'C0392B'), None, PCT, hor='center')
    put(ws, r, 7, '=IF($H%d="","",IFERROR(INDEX(%s,%s),""))' % (r, R_AC_DV, m), FCALC, None, MON, hor='center')

bloque(45, '4 · LOS 12 PRODUCTOS CON MAYOR DIFERENCIA SIN EXPLICAR')
cabecera(ws, 46, ['#', 'AREA', 'PRODUCTO', 'ELABORADO', 'DESCARTE', 'DIF. SIN EXPLICAR', '% S/ELAB.', 'k'], 22)
for n in range(1, 13):
    r = 46 + n
    put(ws, r, 1, n, fnt(9, color=MUTED), BANDFILL if n % 2 else None, hor='center')
    put(ws, r, 8, '=IFERROR(LARGE(%s,%d),"")' % (R_BC_K, n), fnt(8, color=MUTED), None, NUM)
    m = 'MATCH($H%d,%s,0)' % (r, R_BC_K)
    put(ws, r, 2, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_AR, m), fnt(9), BANDFILL if n % 2 else None)
    put(ws, r, 3, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_PR, m), fnt(10, True), BANDFILL if n % 2 else None)
    put(ws, r, 4, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_ELA, m), FCALC, None, NUM0, hor='center')
    put(ws, r, 5, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_DES, m), FCALC, None, NUM0, hor='center')
    put(ws, r, 6, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_DIF, m), fnt(10, True), None, NUM0, hor='center')
    put(ws, r, 7, '=IF($H%d="","",INDEX(%s,%s))' % (r, R_BC_PCT, m), fnt(10, True, 'C0392B'), None, PCT, hor='center')

bloque(60, '5 · DESCARTES POR MOTIVO  (por que se tira)')
cabecera(ws, 61, ['', 'MOTIVO', '', 'CANTIDAD', 'LINEAS', '', '', ''], 20)
for i, m in enumerate(MOTIVOS):
    r = 62 + i
    put(ws, r, 2, m, fnt(10), BANDFILL if i % 2 else None)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    put(ws, r, 3, None, fill=BANDFILL if i % 2 else None)
    put(ws, r, 4, '=SUMIF(%s,$B%d,%s)' % (R_DES_MOT, r, R_DES_CANT), fnt(10, True), None, NUM, hor='center')
    put(ws, r, 5, '=COUNTIF(%s,$B%d)' % (R_DES_MOT, r), fnt(10), None, NUM0, hor='center')
    for c in (6, 7, 8): put(ws, r, c, None, box=False)

bloque(72, '6 · DESCARTES Y ALERTAS POR AREA')
cabecera(ws, 73, ['', 'AREA', '', 'DESCARTE\n(cantidad)', 'LINEAS DE\nDESCARTE', 'INSUMOS\nCON ALERTA', 'PRODUCTOS\nCON ALERTA', ''], 30)
for i, a in enumerate(AREAS):
    r = 74 + i
    put(ws, r, 2, ALAB[a['key']], fnt(10, True), BANDFILL if i % 2 else None)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    put(ws, r, 3, None, fill=BANDFILL if i % 2 else None)
    put(ws, r, 4, '=SUMIF(%s,$B%d,%s)' % (R_DES_AR, r, R_DES_CANT), fnt(10, True), None, NUM, hor='center')
    put(ws, r, 5, '=COUNTIF(%s,$B%d)' % (R_DES_AR, r), fnt(10), None, NUM0, hor='center')
    put(ws, r, 6, '=COUNTIFS(%s,$B%d,%s,"<>OK",%s,"<>")' % (R_AC_AR, r, R_AC_ALE, R_AC_ALE), fnt(10), None, NUM0, hor='center')
    put(ws, r, 7, '=COUNTIFS(%s,$B%d,%s,"<>OK",%s,"<>")' % (R_BC_AR, r, R_BC_ALE, R_BC_ALE), fnt(10), None, NUM0, hor='center')
    put(ws, r, 8, None, box=False)
ws.column_dimensions['H'].hidden = True
ws.freeze_panes = 'A5'
imprimir(ws, 7, 80, landscape=False, repeat='1:3')

# =========================================================================
# Hoja INSTRUCTIVO
# =========================================================================
ws = SH['Instructivo']
anchos(ws, [3, 30, 22, 22, 22, 22, 18, 3])
titulo(ws, 8, 'CONTROL DE STOCK SEMANAL — COMO USAR ESTE LIBRO',
       'Candela Cafe & Patisserie  ·  Prueba de 1 semana  ·  6 areas de produccion, %d insumos, %d productos' % (NINS, sum(len(v) for v in PROD.values())),
       'Leer esta hoja antes de imprimir nada. Toma 5 minutos y evita una semana perdida.')

GUIA = [
 ('h', 'QUE RESPONDE ESTE LIBRO'),
 ('p', 'Cuanto insumo hay, cuanto se consumio de verdad en la semana, y si ese consumo se explica '
       'con lo que se elaboro. Lo que no se explica queda a la vista, separado en cuatro causas posibles: '
       'mal uso del insumo (rinde menos de lo que deberia), consumo del personal sin registrar, '
       'insumo que desaparece, y producto que se quema o se tira.'),
 ('h', 'ANTES DE ARRANCAR  (el domingo previo, 1 hora)'),
 ('li', '1. Hoja CONFIG: poner la fecha del lunes de arranque y el nombre del jefe de cada area. Todas las fechas del libro se completan solas.'),
 ('li', '2. Hoja INSUMOS: revisar UNIDAD BASE y CONTENIDO POR BULTO de los %d insumos. Los puse inferidos del nombre '
        '(ej. "Harina 000 x25kg" -> 25 kg por bulto) pero hay que confirmarlos contra el envase real. Si esto esta mal, todo el analisis sale mal.' % NINS),
 ('li', '3. Hoja INSUMOS, columna CRITICO: vienen %d marcados como SI (los caros y los de mayor riesgo). Esos son los unicos que se cuentan todos los dias. Sumar o quitar los que hagan falta.' % len(CRIT)),
 ('li', '4. Imprimir: la planilla 1 (una por area, usando el filtro de AREA), la planilla 2, la planilla 3 (una por area) y la planilla 4 de cada area.'),
 ('h', 'DURANTE LA SEMANA'),
 ('li', 'LUNES TEMPRANO, antes de producir — PLANILLA 1 (Inventario inicial): cada jefe de area cuenta TODOS sus insumos. '
        'Bultos cerrados + lo que hay suelto del envase abierto. La planilla calcula el total.'),
 ('li', 'TODOS LOS DIAS, cuando entra mercaderia — PLANILLA 2 (Ingresos): quien recibe anota area, insumo, cantidad y numero de remito. '
        'En el momento, no despues. Un ingreso no anotado aparece al final como un faltante que no existe.'),
 ('li', 'TODOS LOS DIAS, al cierre de la jornada — PLANILLA 3 (Conteo diario): cada jefe cuenta solo sus insumos criticos (unos 10). Son 5 minutos.'),
 ('li', 'TODOS LOS DIAS, mientras se produce — PLANILLA 4 (Produccion diaria): el jefe de area anota lo ELABORADO y lo DESCARTADO de cada producto. '
        'Una planilla por area, ya con los productos del catalogo impresos.'),
 ('li', 'TODOS LOS DIAS, cada vez que se tira algo — PLANILLA 5 (Descartes): una linea por descarte, con el MOTIVO. '
        'Esta es la planilla que despues justifica los faltantes: lo que no se anota aca se lee como perdida sin explicacion.'),
 ('li', 'DOMINGO AL CIERRE — PLANILLA 6 (Inventario final): mismo conteo que el lunes, mismo criterio, y si se puede la misma persona.'),
 ('h', 'DESPUES DE LA SEMANA  (carga de escritorio, 2 a 3 horas)'),
 ('li', 'HOJA 7 (Facturas): cargar las facturas de compra de la semana. La cantidad va en la unidad base (kg, u, L), no en bultos. '
        'El importe permite valorizar todo en pesos y llena el costo unitario.'),
 ('li', 'HOJA 8a (Fudo Ventas): pegar el export de ventas por producto y mapear cada linea al producto y area de Candela.'),
 ('li', 'HOJA 8b (Fudo Stock): pegar el reporte de existencias/movimientos y mapear cada insumo.'),
 ('li', 'HOJA 8c (Fudo Consumos): cargar consumos internos y cortesias. Sin esto, el consumo legitimo del personal se confunde con robo.'),
 ('h', 'COMO LEER LOS RESULTADOS'),
 ('p', 'Arranque siempre por el TABLERO, bloque 1: si la carga esta incompleta, los numeros no significan nada.'),
 ('li', 'ANALISIS A (Consumo de insumos): CONSUMO REAL = inicial + ingresos - final. '
        'Si el consumo real supera al de Fudo, ese insumo rinde menos de lo que deberia: mal uso, desperdicio o faltante.'),
 ('li', 'ANALISIS B (Conciliacion de productos): ELABORADO - VENDIDO - CONSUMO INTERNO - DESCARTE. '
        'Lo que sobra es salida sin explicar: consumo del personal no registrado, o producto que se fue sin pasar por caja.'),
 ('li', 'ANALISIS C (Consumo diario): ubica EL DIA del desvio. Un pico atipico en un solo dia apunta a un evento puntual; '
        'un desvio parejo toda la semana apunta a una receta mal calculada o a un habito.'),
 ('h', 'COMO INTERPRETAR CADA SENAL'),
 ('t', ['SENAL EN LA PLANILLA', 'LO QUE PROBABLEMENTE SIGNIFICA', 'QUE HACER']),
 ('r', ['Consumo real muy por encima del de Fudo, parejo toda la semana',
        'La receta rinde menos de lo que dice la ficha tecnica, o se desperdicia en el proceso',
        'Pesar una produccion completa y recalcular la ficha']),
 ('r', ['Consumo real alto pero concentrado en 1 o 2 dias',
        'Un evento puntual: una tanda perdida, un derrame, o un retiro no registrado',
        'Cruzar con la planilla 5 y con quien estuvo de turno ese dia']),
 ('r', ['Consumo negativo (el stock final es mayor al posible)',
        'Entro mercaderia que nadie anoto en la planilla 2',
        'Buscar el remito y reforzar el registro de ingresos']),
 ('r', ['Diferencia sin explicar positiva y grande en un producto',
        'Salio producto sin venderse ni registrarse: consumo del personal o faltante',
        'Comparar con Fudo consumos internos y con el sobrante en exhibicion']),
 ('r', ['Descarte declarado muy por debajo de lo que se ve en el piso',
        'No se esta anotando lo que se tira',
        'Insistir en la planilla 5: sin motivo cargado, todo se lee como faltante']),
 ('r', ['Lo recibido no coincide con la factura',
        'Se factura mas de lo que entra, o se recibe sin controlar',
        'Reclamar al proveedor y pesar en la recepcion']),
 ('h', 'LIMITES DE ESTE METODO  (para no sacar conclusiones de mas)'),
 ('warn', 'NO se lleva stock de producto terminado. La diferencia del analisis B incluye lo que quedo en exhibicion '
          'al cierre del domingo. Lea la tendencia y los casos grandes, no el numero exacto de cada producto.'),
 ('warn', 'Una sola semana no distingue una causa estable de una casualidad. Sirve para senalar donde mirar, no para acusar a nadie.'),
 ('warn', 'La calidad del resultado depende del conteo. Si el inventario inicial se hace "a ojo", el libro devuelve numeros prolijos y falsos.'),
 ('warn', 'Cocina San Luis no tiene productos en el catalogo de la app: su planilla 4 viene en blanco para escribirlos a mano.'),
 ('h', 'REGLA DE ORO'),
 ('p', 'Solo se escribe en las celdas AMARILLAS. Todo lo que esta en negro o en verde es formula: si se pisa, se rompe el analisis.'),
]

r = 5
for tipo, txt in GUIA:
    if tipo == 'h':
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=txt)
        c.font = fnt(11, True, 'FFFFFF'); c.fill = SUBFILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[r].height = 22
        r += 1
    elif tipo in ('p', 'li', 'warn'):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=('•  ' + txt) if tipo == 'li' else txt)
        c.font = fnt(10, tipo == 'warn', 'C0392B' if tipo == 'warn' else TXT)
        if tipo == 'warn':
            c.fill = PatternFill('solid', fgColor=DANGBG)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        ws.row_dimensions[r].height = max(15, 13 * (len(txt) // 118 + 1))
        r += 1
    elif tipo == 't':
        for i, h in enumerate(txt):
            put(ws, r, 2 + i * 2, h, fnt(9, True, 'FFFFFF'), SUBFILL, hor='center', wrap=True)
            ws.merge_cells(start_row=r, start_column=2 + i * 2, end_row=r, end_column=3 + i * 2)
            put(ws, r, 3 + i * 2, None, fill=SUBFILL)
        ws.row_dimensions[r].height = 24
        r += 1
    elif tipo == 'r':
        for i, v in enumerate(txt):
            put(ws, r, 2 + i * 2, v, fnt(9), None, None, wrap=True, ver='top')
            ws.merge_cells(start_row=r, start_column=2 + i * 2, end_row=r, end_column=3 + i * 2)
            put(ws, r, 3 + i * 2, None)
        ws.row_dimensions[r].height = 30
        r += 1
ws.freeze_panes = 'A5'
imprimir(ws, 7, r - 1, landscape=False, repeat='1:3')

wb.properties.title = 'Control de Stock Semanal - Candela Cafe & Patisserie'
wb.properties.creator = 'Candela'
OUT = SALIDA
wb.save(OUT)
print('guardado:', OUT, '| hojas:', len(wb.sheetnames))
